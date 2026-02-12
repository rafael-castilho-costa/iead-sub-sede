#!/usr/bin/env python3
"""Import birthdays from an Excel workbook into JSON used by Home page.

Usage:
  python scripts/import-aniversariantes.py \
    --input "C:\\Users\\...\\Aniversariantes Sub Sede - 2026.xltx.xlsx" \
    --output "src/assets/data/aniversariantes.json"

The script expects each month sheet to have:
- name in column B
- date in column D
- rows starting at 4
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl

MONTH_BY_SHEET = {
    "Janeiro": 1,
    "Fevereiro": 2,
    "Março": 3,
    "Abril": 4,
    "Maio": 5,
    "Junho": 6,
    "Julho": 7,
    "Agosto": 8,
    "Setembro": 9,
    "Outubro": 10,
    "Novembro": 11,
    "Dezembro": 12,
}


@dataclass(frozen=True)
class Birthday:
    name: str
    month: int
    day: int

    @property
    def mmdd(self) -> str:
        return f"{self.month:02d}-{self.day:02d}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import birthdays from Excel to JSON")
    parser.add_argument("--input", required=True, help="Path to XLSX/XLTX file")
    parser.add_argument(
        "--output",
        default="src/assets/data/aniversariantes.json",
        help="Output JSON path (default: src/assets/data/aniversariantes.json)",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=4,
        help="Start row for data in each sheet (default: 4)",
    )
    parser.add_argument(
        "--dedupe",
        action="store_true",
        default=True,
        help="Remove exact duplicates by name+date (default: enabled)",
    )
    parser.add_argument(
        "--strict-sheet-month",
        action="store_true",
        help=(
            "Force month based on sheet name when date month differs. "
            "Useful when spreadsheet has month typing errors."
        ),
    )
    return parser.parse_args()


def parse_date(value: object) -> tuple[int, int] | None:
    if isinstance(value, (datetime, date)):
        return value.month, value.day

    if isinstance(value, (int, float)):
        try:
            dt = openpyxl.utils.datetime.from_excel(value)
            return dt.month, dt.day
        except Exception:
            return None

    if isinstance(value, str) and value.strip():
        raw = value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(raw, fmt)
                return dt.month, dt.day
            except ValueError:
                continue

    return None


def iter_birthdays(
    workbook_path: Path,
    start_row: int,
    strict_sheet_month: bool,
) -> tuple[list[Birthday], list[str]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    entries: list[Birthday] = []
    warnings: list[str] = []

    for ws in wb.worksheets:
        expected_month = MONTH_BY_SHEET.get(ws.title)

        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=5):
            raw_name = row[1].value  # Column B
            raw_date = row[3].value  # Column D

            if raw_name is None:
                continue

            name = str(raw_name).strip()
            if not name:
                continue

            parsed = parse_date(raw_date)
            if parsed is None:
                warnings.append(
                    f"[WARN] {ws.title} row {row[1].row}: invalid date for '{name}' -> {raw_date!r}"
                )
                continue

            month, day = parsed

            if expected_month and month != expected_month:
                warnings.append(
                    f"[WARN] {ws.title} row {row[1].row}: month mismatch for '{name}' "
                    f"({month:02d}-{day:02d})"
                )
                if strict_sheet_month:
                    month = expected_month

            entries.append(Birthday(name=name, month=month, day=day))

    return entries, warnings


def dedupe_birthdays(items: Iterable[Birthday]) -> list[Birthday]:
    seen: set[tuple[str, str]] = set()
    unique: list[Birthday] = []
    for item in items:
        key = (item.name.casefold(), item.mmdd)
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
    return unique


def to_payload(items: Iterable[Birthday]) -> list[dict[str, str]]:
    ordered = sorted(items, key=lambda i: (i.mmdd, i.name.casefold()))
    return [{"nome": item.name, "data": item.mmdd} for item in ordered]


def main() -> int:
    args = parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    entries, warnings = iter_birthdays(
        workbook_path=input_path,
        start_row=args.start_row,
        strict_sheet_month=args.strict_sheet_month,
    )

    if args.dedupe:
        entries = dedupe_birthdays(entries)

    payload = to_payload(entries)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Imported {len(payload)} records -> {output_path}")
    if warnings:
        print(f"Warnings: {len(warnings)}")
        for line in warnings:
            print(line)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
